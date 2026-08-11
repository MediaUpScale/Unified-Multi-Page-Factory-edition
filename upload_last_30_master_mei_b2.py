#!/usr/bin/env python3
"""Upload last-30 master_mei clips from postplan_20260803 to B2 and patch MEDIA URLs."""
from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(_ROOT))
load_dotenv(_ROOT / ".env", override=False)

from avatar_engine.b2_client import B2VideoUploader  # noqa: E402

_PAGE = "master_mei"
_PP = _ROOT / "outputs" / _PAGE / "postplanner" / "postplan_20260803_061313.xlsx"
_BULK = _ROOT / "outputs" / _PAGE / "automated_bulk_posts_import.xlsx"
_LIB = _ROOT / "outputs" / _PAGE / "content_library.json"


def main() -> int:
    if not _PP.is_file():
        print(f"MISSING postplan: {_PP}")
        return 1

    wb = load_workbook(_PP)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
    paths: list[Path] = []
    for row in rows:
        media = row[2].value
        if not media:
            continue
        p = Path(str(media))
        if p.is_file() and p.suffix.lower() == ".mp4":
            paths.append(p)
        else:
            print(f"SKIP missing local: {media}")

    print(f"Uploading {len(paths)} clips …")
    b2 = B2VideoUploader()
    url_by_name: dict[str, str] = {}
    ok = fail = 0
    for i, p in enumerate(paths, 1):
        print(f"\n[{i}/{len(paths)}] {p.name} ({p.stat().st_size / 1e6:.1f} MB)")
        try:
            url = b2.upload(p, content_type="video/mp4")
            if not url or "backblazeb2.com" not in url:
                raise RuntimeError(f"Non-B2 URL returned: {url!r}")
            if not b2._object_exists(p.name):
                raise RuntimeError(f"Post-upload HEAD miss for {p.name}")
            url_by_name[p.name] = url
            ok += 1
            print(f"  OK {url}")
        except Exception as exc:
            fail += 1
            print(f"  FAIL {type(exc).__name__}: {exc}")

    if not url_by_name:
        print("No successful uploads — aborting Excel patch.")
        return 2

    # Patch primary postplan
    bak = _PP.with_suffix(_PP.suffix + ".pre_b2_20260809.bak")
    if not bak.is_file():
        shutil.copy2(_PP, bak)
    patched = 0
    for row in rows:
        media = row[2].value
        if not media:
            continue
        name = Path(str(media)).name
        if name in url_by_name:
            row[2].value = url_by_name[name]
            patched += 1
    wb.save(_PP)
    print(f"\nPatched {_PP.name}: {patched} MEDIA URLs")

    # Patch bulk import if present
    if _BULK.is_file():
        bb = load_workbook(_BULK)
        bws = bb.active
        headers = [c.value for c in next(bws.iter_rows(min_row=1, max_row=1))]
        try:
            media_col = next(i for i, h in enumerate(headers) if h and "MEDIA" in str(h).upper()) + 1
        except StopIteration:
            media_col = 3
        b_patched = 0
        for brow in bws.iter_rows(min_row=2, max_row=bws.max_row):
            cell = brow[media_col - 1]
            if not cell.value:
                continue
            name = Path(str(cell.value)).name
            if name in url_by_name:
                cell.value = url_by_name[name]
                b_patched += 1
        if b_patched:
            bb.save(_BULK)
            print(f"Patched {_BULK.name}: {b_patched} MEDIA URLs")

    # Patch content_library video_path → keep local path; add/update imgbb_url-like public url field if used as media
    if _LIB.is_file():
        lib = json.loads(_LIB.read_text(encoding="utf-8"))
        lib_n = 0
        for entry in lib:
            vp = entry.get("video_path") or ""
            name = Path(str(vp)).name
            if name in url_by_name:
                # Preserve local video_path; store public URL in imgbb_url (legacy media URL field)
                entry["imgbb_url"] = url_by_name[name]
                entry["b2_url"] = url_by_name[name]
                lib_n += 1
        if lib_n:
            _LIB.write_text(json.dumps(lib, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"Patched content_library.json: {lib_n} entries")

    print(f"\nDONE ok={ok} fail={fail} urls={len(url_by_name)}")
    return 0 if fail == 0 else 3


if __name__ == "__main__":
    raise SystemExit(main())
