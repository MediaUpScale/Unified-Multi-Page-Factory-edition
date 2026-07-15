#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
_repair_postplanner_b2.py
=========================
One-shot repair script:
  1. Reads deep psychological captions from postplan_20260606_231538.xlsx
  2. Reads 57 B2 media URLs from PostPlanner_B2_20260608_003444.xlsx
  3. Cycles captions to fill all 57 video rows (with repetition noted)
  4. Appends the signature hashtag block to every caption
  5. Preserves the 4-row PostPlanner template header (rows 1-4 in the sheet)
  6. Clears all rows below the header and writes clean data from row 5 onward
  7. Moves URLs to the correct CONTENT: MEDIA column (col D, index 3)
  8. Sets POST TYPE = VIDEO for every row
  9. Overwrites PostPlanner_B2_20260608_003444.xlsx in-place
"""
from __future__ import annotations

import sys
from itertools import cycle
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
ROOT        = Path(__file__).parent
POSTPLAN_DIR = ROOT / "outputs/wonder_feed/postplanner"

CAPTION_SRC  = POSTPLAN_DIR / "postplan_20260606_231538.xlsx"
TARGET_FILE  = POSTPLAN_DIR / "PostPlanner_B2_20260608_003444.xlsx"

SIGNATURE_TAGS = "💙 #Relationships #EmotionalIntelligence #Psychology #WonderFeed"

# PostPlanner 6-column schema (template header rows 1–4)
#  col 1: ## POSTING TIME
#  col 2: CAPTION
#  col 3: CONTENT: LINK   ← leave blank for video posts
#  col 4: CONTENT: MEDIA  ← B2 public URL goes here
#  col 5: POST TYPE
#  col 6: PIN TITLE
TEMPLATE_HEADER_ROWS = 4   # rows 1-4 in the sheet (1-indexed) are kept untouched
DATA_START_ROW       = 5   # first writable data row (1-indexed, openpyxl)

# ---------------------------------------------------------------------------
# Step 1 — Read captions
# ---------------------------------------------------------------------------
try:
    import pandas as pd
except ImportError:
    print("[ERROR] pandas not installed.  Run: pip install pandas openpyxl")
    sys.exit(1)

print(f"\n[1] Reading captions from: {CAPTION_SRC.name}")
src = pd.read_excel(str(CAPTION_SRC), header=None)
# Row 0 is the header ("DATE / TIME", "CAPTION", "MEDIA URL")
# Captions live in column index 1, rows 1 onward
raw_captions = [
    str(v).strip()
    for v in src.iloc[1:, 1].tolist()
    if str(v).strip() not in ("", "nan", "CAPTION")
]
print(f"    Captions found: {len(raw_captions)}")
for i, c in enumerate(raw_captions, 1):
    print(f"    {i:>2}. {c[:90]}")

if not raw_captions:
    print("[ERROR] No captions found in source file.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Step 2 — Read B2 URLs + posting times from target file
# ---------------------------------------------------------------------------
print(f"\n[2] Reading B2 data from: {TARGET_FILE.name}")
df_target = pd.read_excel(str(TARGET_FILE), header=None)

# Rows where column 2 (CONTENT: LINK — the wrong column our script used) has a B2 URL
b2_mask = df_target.iloc[:, 2].astype(str).str.contains("backblazeb2.com", na=False)
b2_rows  = df_target[b2_mask]
print(f"    B2 URL rows found: {len(b2_rows)}  (pandas idx {b2_rows.index[0]}–{b2_rows.index[-1]})")

posting_times = b2_rows.iloc[:, 0].astype(str).tolist()
b2_urls       = b2_rows.iloc[:, 2].astype(str).tolist()   # currently in CONTENT: LINK

if not b2_urls:
    print("[ERROR] No B2 URLs found in target file.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Step 3 — Build aligned records (cycle captions if fewer than URLs)
# ---------------------------------------------------------------------------
n_videos = len(b2_urls)
n_caps   = len(raw_captions)

if n_caps < n_videos:
    print(f"\n[3] Captions ({n_caps}) < videos ({n_videos}) — cycling captions to fill all rows.")
else:
    print(f"\n[3] Aligning {n_caps} captions → {n_videos} videos.")

cap_cycle = cycle(raw_captions)
records: list[dict] = []
for idx, (url, time_str) in enumerate(zip(b2_urls, posting_times)):
    caption_raw = next(cap_cycle)
    caption_full = f"{caption_raw}\n\n{SIGNATURE_TAGS}"
    records.append({
        "posting_time": "" if time_str in ("nan", "") else time_str,
        "caption":      caption_full,
        "media_url":    url,
    })

print(f"    Records to write: {len(records)}")

# ---------------------------------------------------------------------------
# Step 4 — Rewrite the target workbook in-place
# ---------------------------------------------------------------------------
print(f"\n[4] Rewriting: {TARGET_FILE.name}")

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl not installed.  Run: pip install openpyxl")
    sys.exit(1)

wb = load_workbook(str(TARGET_FILE))
ws = wb.active

# Delete every row from DATA_START_ROW onward (openpyxl 1-indexed)
# We delete from the bottom up to avoid index shifting.
max_r = ws.max_row
if max_r >= DATA_START_ROW:
    ws.delete_rows(DATA_START_ROW, max_r - DATA_START_ROW + 1)
    print(f"    Cleared rows {DATA_START_ROW}–{max_r} ({max_r - DATA_START_ROW + 1} rows removed)")

# Write clean records
for i, rec in enumerate(records):
    row = DATA_START_ROW + i
    ws.cell(row=row, column=1, value=rec["posting_time"])  # ## POSTING TIME
    ws.cell(row=row, column=2, value=rec["caption"])        # CAPTION
    ws.cell(row=row, column=3, value="")                    # CONTENT: LINK  (blank)
    ws.cell(row=row, column=4, value=rec["media_url"])      # CONTENT: MEDIA ← correct column
    ws.cell(row=row, column=5, value="VIDEO")               # POST TYPE
    ws.cell(row=row, column=6, value="")                    # PIN TITLE

wb.save(str(TARGET_FILE))

print(f"\n[DONE] {len(records)} rows written → {TARGET_FILE}")
print(f"       Data starts at sheet row {DATA_START_ROW}, URLs now in CONTENT: MEDIA (col D).")
print(f"       Signature tags appended to every caption.")
