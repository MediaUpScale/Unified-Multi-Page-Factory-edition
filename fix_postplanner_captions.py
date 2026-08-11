"""fix_postplanner_captions.py — Clean raw LLM JSON out of PostPlanner CAPTION cells.

Reads one or more postplan_*.xlsx files, parses any markdown-fenced / JSON
caption blobs into plain ``caption_body`` text, and writes the cleaned values
back in-place.

Usage
-----
    python fix_postplanner_captions.py
    python fix_postplanner_captions.py path/to/postplan_....xlsx
    python fix_postplanner_captions.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from openpyxl import load_workbook  # noqa: E402

from avatar_engine.post_planner import extract_clean_caption  # noqa: E402

_DEFAULT_TARGET = (
    _SCRIPT_DIR
    / "outputs"
    / "ancient_knowledge"
    / "postplanner"
    / "postplan_20260802_212650.xlsx"
)
_COL_CAPTION = 2


def _looks_dirty(text: str) -> bool:
    s = (text or "").strip()
    if not s:
        return False
    if s.startswith("```") or s.startswith("{"):
        return True
    head = s[:120]
    return "caption_body" in head or '"caption"' in head


def fix_workbook(path: Path, *, dry_run: bool = False) -> tuple[int, int]:
    """Clean CAPTION cells in *path*. Returns ``(rows_scanned, rows_fixed)``."""
    path = Path(path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)

    wb = load_workbook(path)
    try:
        ws = wb.active
        scanned = 0
        fixed = 0
        for row in range(2, (ws.max_row or 1) + 1):
            cell = ws.cell(row=row, column=_COL_CAPTION)
            raw = cell.value
            if raw is None:
                continue
            scanned += 1
            raw_s = str(raw)
            cleaned = extract_clean_caption(raw_s)
            if cleaned == raw_s.strip() and not _looks_dirty(raw_s):
                continue
            if cleaned == raw_s:
                continue
            print(
                f"  row {row}: {len(raw_s)} -> {len(cleaned)} chars "
                f"| start={cleaned[:70]!r}"
            )
            if not dry_run:
                cell.value = cleaned
            fixed += 1

        if not dry_run and fixed:
            wb.save(path)
    finally:
        if callable(getattr(wb, "close", None)):
            wb.close()

    return scanned, fixed


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Strip raw LLM JSON/markdown from PostPlanner CAPTION cells."
    )
    parser.add_argument(
        "paths",
        nargs="*",
        type=Path,
        help="XLSX file(s) to fix (default: latest corrupted ancient_knowledge postplan)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without writing",
    )
    args = parser.parse_args()

    targets = [Path(p) for p in args.paths] if args.paths else [_DEFAULT_TARGET]
    total_fixed = 0
    for target in targets:
        print(f"{'[DRY-RUN] ' if args.dry_run else ''}Fixing: {target}")
        scanned, fixed = fix_workbook(target, dry_run=args.dry_run)
        print(f"  scanned={scanned} fixed={fixed}")
        total_fixed += fixed

    print(f"Done. Total captions cleaned: {total_fixed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
