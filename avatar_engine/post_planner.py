# -*- coding: utf-8 -*-
"""
PostPlanner -- Excel (.xlsx) bulk-import writer for the content pipeline.

Column layout (3-column PostPlanner format):
  A: DATE / TIME  -- blank when PUBLISHING_SCHEDULE is None
  B: CAPTION      -- humanized caption with dynamic CTA woven in
  C: MEDIA URL    -- B2 public .mp4 URL for video/reel posts;
                     ImgBB display_url for static image posts

CSV output is retired; all bulk outputs are .xlsx only.
"""
from __future__ import annotations

import json
import logging
import os
import re
import time
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

_LOG = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions (3-column layout)
# ---------------------------------------------------------------------------

_COL_DATETIME = 1   # A
_COL_CAPTION = 2    # B
_COL_MEDIA = 3      # C

_FALLBACK_HEADERS: list[str] = [
    "DATE / TIME",
    "CAPTION",
    "MEDIA URL",
]

# openpyxl / Spreadsheet ML rejects certain control chars.
_EXCEL_ILLEGAL_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_CAPTION_CELL_MAX_LEN = 32_000
_URL_CELL_MAX_LEN = 2048

# Keys the LLM may use for the human-readable post body.
_CAPTION_JSON_KEYS: tuple[str, ...] = (
    "caption_body",
    "caption",
    "description",
    "text",
    "post_caption",
)
_FENCE_BLOCK_RE = re.compile(
    r"^```(?:json|JSON)?\s*\n?(.*?)\n?```\s*$",
    re.DOTALL,
)
_FENCE_START_RE = re.compile(r"^```(?:json|JSON)?\s*\n?", re.IGNORECASE)
_FENCE_END_RE = re.compile(r"\n?```\s*$")
_JSON_KEY_ARTIFACT_RE = re.compile(
    r'"?\b(caption_body|caption|description|image_text_overlay|visual_subject|'
    r'quote_text|seo_title|text|post_caption)\b"?\s*:\s*',
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Caption sanitation (LLM JSON / markdown → plain text)
# ---------------------------------------------------------------------------


def extract_clean_caption(raw: Any) -> str:
    """
    Extract plain post-caption text from LLM output.

    Handles:
      * dict payloads with ``caption_body`` / ``caption`` / etc.
      * markdown-fenced JSON (```json ... ```)
      * serialized JSON objects (valid or with literal newlines inside strings)
      * already-clean plain text (returned unchanged)

    Always returns a plain ``str`` suitable for PostPlanner CAPTION cells.
    """
    if raw is None:
        return ""

    if isinstance(raw, dict):
        for key in _CAPTION_JSON_KEYS:
            val = raw.get(key)
            if isinstance(val, str) and val.strip():
                return val.strip()
        return ""

    if isinstance(raw, (list, tuple)):
        return " ".join(str(v) for v in raw).strip()

    text = str(raw).strip()
    if not text:
        return ""

    fence = _FENCE_BLOCK_RE.match(text)
    if fence:
        text = fence.group(1).strip()
    elif text.startswith("```"):
        text = _FENCE_START_RE.sub("", text)
        text = _FENCE_END_RE.sub("", text).strip()

    if text[:1] in "{[":
        try:
            parsed = json.loads(text)
            extracted = extract_clean_caption(parsed)
            if extracted:
                return extracted
        except (json.JSONDecodeError, TypeError, ValueError):
            pass

        # Salvage caption_body from invalid JSON (common: unescaped newlines
        # inside string values → json.loads raises Invalid control character).
        for key in _CAPTION_JSON_KEYS:
            for pat in (
                rf'"{key}"\s*:\s*"(.*?)"\s*,?\s*\n\s*"',
                rf'"{key}"\s*:\s*"(.*?)"\s*\n?\s*[}}\]]',
            ):
                match = re.search(pat, text, re.DOTALL)
                if match:
                    val = match.group(1)
                    val = (
                        val.replace("\\n", "\n")
                        .replace("\\t", "\t")
                        .replace('\\"', '"')
                        .replace("\\\\", "\\")
                    )
                    return val.strip()

        # Last resort: drop braces/brackets so we don't ship raw JSON wrappers.
        text = text.strip("{}[]").strip()
        text = _JSON_KEY_ARTIFACT_RE.sub("", text).strip()

    return text


def _sanitize_caption_for_export(value: Any) -> str:
    """Parse LLM JSON/markdown then strip Excel-illegal control characters."""
    return _sanitize_excel_cell_text(extract_clean_caption(value))


# ---------------------------------------------------------------------------
# Scheduling helpers
# ---------------------------------------------------------------------------


def scheduled_bulk_post_display(
    *,
    variant_index: int = 0,
    interval_minutes: int | None = None,
    stagger_minutes: int = 7,
) -> str:
    """
    Return a ``MM/DD/YYYY HH:MM`` posting slot, or an empty string when no
    schedule is configured (Null Schedule Protocol).

    Modes
    -----
    1. ``PUBLISHING_SCHEDULE = None`` in config  ->  returns ``""`` (blank cell).
    2. ``interval_minutes`` supplied (or from ``config.PUBLISHING_INTERVAL_MINUTES``):
       Slot = now + interval_minutes * variant_index.
    3. Legacy fallback (interval_minutes explicitly given without config):
       Slot = now + 24h + stagger_minutes * variant_index.
    """
    try:
        import config as _cfg  # noqa: PLC0415
        resolved_interval = (
            interval_minutes
            if interval_minutes is not None
            else _cfg.PUBLISHING_INTERVAL_MINUTES
        )
    except Exception:  # noqa: BLE001
        resolved_interval = interval_minutes

    # Null Schedule Protocol: leave the cell blank
    if resolved_interval is None:
        return ""

    anchor = datetime.now().astimezone()
    dt = anchor + timedelta(minutes=resolved_interval * int(variant_index))
    return dt.strftime("%m/%d/%Y %H:%M")


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _sanitize_excel_cell_text(value: Any, *, max_len: int = _CAPTION_CELL_MAX_LEN) -> str:
    """Drop illegal XML/control characters and truncate for a single worksheet cell."""
    s = "" if value is None else str(value)
    s = _EXCEL_ILLEGAL_CTRL.sub("", s)
    return s[:max_len] if len(s) > max_len else s


def _copy_cell_style(src: Any, dst: Any) -> None:
    try:
        if getattr(src, "font", None) and src.font:
            dst.font = copy(src.font)
        if getattr(src, "border", None) and src.border:
            dst.border = copy(src.border)
        if getattr(src, "fill", None) and src.fill:
            dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        if getattr(src, "protection", None) and src.protection:
            dst.protection = copy(src.protection)
        if getattr(src, "alignment", None) and src.alignment:
            dst.alignment = copy(src.alignment)
    except (TypeError, ValueError):
        pass


# ---------------------------------------------------------------------------
# Workbook bootstrap
# ---------------------------------------------------------------------------


def _create_fallback_workbook(destination_path: Path) -> None:
    destination_path = Path(destination_path).expanduser().resolve()
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    for col_idx, title in enumerate(_FALLBACK_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=title)
    _save_workbook(wb, destination_path)


def _clone_header_sheet_from_template(template_path: Path, destination_path: Path) -> None:
    """Clone the header row from a template workbook; used for branded sheets."""
    template_path = Path(template_path).expanduser().resolve()
    destination_path = Path(destination_path).expanduser().resolve()
    destination_path.parent.mkdir(parents=True, exist_ok=True)

    src_wb = load_workbook(template_path, keep_vba=False, data_only=False)
    src_ws = src_wb.active

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = src_ws.title

    max_col = max(1, int(src_ws.max_column or 1))

    for rng in tuple(src_ws.merged_cells.ranges):
        try:
            if rng.min_row == rng.max_row == 1:
                dst_ws.merge_cells(str(rng))
        except (ValueError, KeyError):
            continue

    for col in range(1, max_col + 1):
        src_cell = src_ws.cell(row=1, column=col)
        dst_cell = dst_ws.cell(row=1, column=col, value=src_cell.value)
        _copy_cell_style(src_cell, dst_cell)

    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        src_dim = src_ws.column_dimensions.get(letter)
        if src_dim is not None and getattr(src_dim, "width", None):
            dst_ws.column_dimensions[letter].width = src_dim.width

    if 1 in src_ws.row_dimensions and getattr(src_ws.row_dimensions[1], "height", None):
        dst_ws.row_dimensions[1].height = src_ws.row_dimensions[1].height

    freeze = getattr(src_ws, "freeze_panes", None)
    if freeze:
        dst_ws.freeze_panes = freeze

    _save_workbook(dst_wb, destination_path)


def _save_workbook(wb: Workbook, destination_path: Path) -> None:
    """
    Persist ``wb`` to an absolute path with mkdir + retry.

    Google Drive / Excel file locks are common; we write a sibling ``.tmp``
    then atomically replace, retrying once on PermissionError/OSError.
    """
    path = Path(destination_path).expanduser().resolve()
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        _LOG.error("PostPlanner mkdir failed | dir=%s | %s", path.parent, exc)
        raise

    tmp = path.with_name(path.name + ".tmp")
    last_exc: BaseException | None = None
    for attempt in range(1, 3):
        try:
            wb.save(tmp)
            os.replace(str(tmp), str(path))
            _LOG.info("PostPlanner saved | %s", path)
            return
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            _LOG.warning(
                "PostPlanner save failed (attempt %s/2) | path=%s | %s",
                attempt, path, exc,
            )
            try:
                if tmp.is_file():
                    tmp.unlink()
            except OSError:
                pass
            # Direct save as fallback (Drive sometimes rejects os.replace).
            try:
                wb.save(path)
                _LOG.info("PostPlanner saved via direct write | %s", path)
                return
            except Exception as direct_exc:  # noqa: BLE001
                last_exc = direct_exc
                _LOG.warning(
                    "PostPlanner direct save also failed (attempt %s/2) | %s",
                    attempt, direct_exc,
                )
            time.sleep(0.6 * attempt)
    _LOG.error("PostPlanner write FAILED | path=%s | %s", path, last_exc)
    raise OSError(f"PostPlanner could not write {path}: {last_exc}") from last_exc


def _ensure_automated_workbook(automation_path: Path, *, template_path: Path | None) -> None:
    path = Path(automation_path).expanduser().resolve()
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        _LOG.error("PostPlanner parent mkdir failed | dir=%s | %s", path.parent, exc)
        raise
    if path.is_file():
        return
    tmpl = Path(template_path).expanduser().resolve() if template_path else None
    if tmpl and tmpl.is_file():
        try:
            _clone_header_sheet_from_template(tmpl, path)
            return
        except Exception as exc:  # noqa: BLE001
            _LOG.warning(
                "PostPlanner template clone failed (%s) — using fallback headers. tmpl=%s dest=%s",
                exc, tmpl, path,
            )
    _create_fallback_workbook(path)


# ---------------------------------------------------------------------------
# Primary write / update functions (3-column layout)
# ---------------------------------------------------------------------------


def append_planner_row(
    path: Path,
    *,
    posting_time: str,
    caption: str,
    media_url: str = "",
    template_path: Path | None = None,
    # Legacy params kept for call-site compatibility -- silently ignored
    url_link: str = "",
    post_type_value: str | None = None,
) -> int:
    """
    Append one data row to the 3-column PostPlanner workbook.

    Column layout: DATE / TIME | CAPTION | MEDIA URL

    When ``posting_time`` is an empty string (Null Schedule Protocol),
    the DATE / TIME cell is written as blank so PostPlanner queues it
    internally.

    Returns the 1-based Excel row index of the new row.
    """
    path = Path(path).expanduser().resolve()
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        _ensure_automated_workbook(path, template_path=template_path)

        wb = load_workbook(path)
        try:
            ws = wb.active
            next_row = max(2, (ws.max_row or 1) + 1)

            # DATE / TIME -- blank when posting_time is ""
            ws.cell(
                row=next_row,
                column=_COL_DATETIME,
                value=_sanitize_excel_cell_text(posting_time, max_len=64) or None,
            )
            ws.cell(
                row=next_row,
                column=_COL_CAPTION,
                value=_sanitize_caption_for_export(caption),
            )
            ws.cell(
                row=next_row,
                column=_COL_MEDIA,
                value=_sanitize_excel_cell_text(media_url, max_len=_URL_CELL_MAX_LEN) or None,
            )

            _save_workbook(wb, path)
        finally:
            if callable(getattr(wb, "close", None)):
                wb.close()
        return next_row
    except Exception as exc:  # noqa: BLE001
        _LOG.error("append_planner_row failed | path=%s | %s", path, exc)
        raise


def update_planner_row(
    path: Path,
    row_index: int,
    *,
    caption: str | None = None,
    media_url: str | None = None,
    posting_time: str | None = None,
    # Legacy params -- silently ignored
    url_link: str | None = None,
    post_type_value: str | None = None,
) -> None:
    """Partial cell updates on an existing workbook row (3-column layout)."""
    path = Path(path).expanduser().resolve()
    row_index = int(row_index)
    if row_index < 2:
        raise ValueError("Excel data rows begin at row_index >= 2 (row 1 is headers).")

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = load_workbook(path)
        try:
            ws = wb.active

            if posting_time is not None:
                ws.cell(
                    row=row_index,
                    column=_COL_DATETIME,
                    value=_sanitize_excel_cell_text(posting_time, max_len=64) or None,
                )
            if caption is not None:
                ws.cell(
                    row=row_index,
                    column=_COL_CAPTION,
                    value=_sanitize_caption_for_export(caption),
                )
            if media_url is not None:
                ws.cell(
                    row=row_index,
                    column=_COL_MEDIA,
                    value=_sanitize_excel_cell_text(media_url, max_len=_URL_CELL_MAX_LEN) or None,
                )

            _save_workbook(wb, path)
        finally:
            if callable(getattr(wb, "close", None)):
                wb.close()
    except Exception as exc:  # noqa: BLE001
        _LOG.error("update_planner_row failed | path=%s row=%s | %s", path, row_index, exc)
        raise


def update_planner_row_caption(path: Path, row_index: int, *, caption: str) -> None:
    """Backward-compatible alias: caption column only."""
    update_planner_row(path, row_index, caption=caption)


# ---------------------------------------------------------------------------
# Per-run timestamped XLSX export (replaces legacy CSV)
# ---------------------------------------------------------------------------


def append_postplanner_xlsx_row(
    postplanner_dir: Path,
    *,
    run_stamp: str,
    posting_time: str,
    caption: str,
    media_url: str = "",
) -> Path:
    """
    Append one row to ``outputs/postplanner/postplan_<run_stamp>.xlsx``.

    Creates the workbook (with header) on first write; appends on subsequent
    calls.  Returns the path to the workbook.

    Column layout: DATE / TIME | CAPTION | MEDIA URL
    """
    postplanner_dir = Path(postplanner_dir).expanduser().resolve()
    try:
        postplanner_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        _LOG.error("PostPlanner dir mkdir failed | dir=%s | %s", postplanner_dir, exc)
        raise

    xlsx_path = postplanner_dir / f"postplan_{run_stamp}.xlsx"
    try:
        _ensure_automated_workbook(xlsx_path, template_path=None)

        wb = load_workbook(xlsx_path)
        try:
            ws = wb.active
            next_row = max(2, (ws.max_row or 1) + 1)

            ws.cell(
                row=next_row,
                column=_COL_DATETIME,
                value=_sanitize_excel_cell_text(posting_time, max_len=64) or None,
            )
            ws.cell(
                row=next_row,
                column=_COL_CAPTION,
                value=_sanitize_caption_for_export(caption),
            )
            ws.cell(
                row=next_row,
                column=_COL_MEDIA,
                value=_sanitize_excel_cell_text(media_url, max_len=_URL_CELL_MAX_LEN) or None,
            )

            _save_workbook(wb, xlsx_path)
        finally:
            if callable(getattr(wb, "close", None)):
                wb.close()
    except Exception as exc:  # noqa: BLE001
        _LOG.error("append_postplanner_xlsx_row failed | path=%s | %s", xlsx_path, exc)
        raise

    return xlsx_path


# ---------------------------------------------------------------------------
# Deprecated: CSV writer kept as no-op to avoid import errors during rollout
# ---------------------------------------------------------------------------


def append_csv_post_row(
    postplanner_dir: Path,
    *,
    run_stamp: str,
    posting_time: str,
    caption: str,
    media_url: str = "",
    url_link: str = "",
    post_type_value: str = "IMAGE",
) -> Path:
    """
    DEPRECATED -- replaced by append_postplanner_xlsx_row.

    Delegates to the XLSX writer so any existing call sites continue to
    produce output; the CSV file is no longer created.
    """
    return append_postplanner_xlsx_row(
        postplanner_dir,
        run_stamp=run_stamp,
        posting_time=posting_time,
        caption=caption,
        media_url=media_url,
    )
