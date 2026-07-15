#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
upload_clips_to_b2.py — Retroactive B2 upload + PostPlanner generator
======================================================================

Scans a directory of already-generated video clips, uploads each one to
Backblaze B2, and writes a PostPlanner XLSX file with columns:

    DATE / TIME  |  CAPTION  |  MEDIA URL  (B2 public URL)

Usage
-----
    # Upload every .mp4 in outputs/wonder_feed/clips/ and generate a planner
    python upload_clips_to_b2.py

    # Specify a custom clips folder and optional caption text
    python upload_clips_to_b2.py --clips-dir path/to/clips --page wonder_feed

    # Dry-run: print what would be uploaded without touching B2
    python upload_clips_to_b2.py --dry-run

    # Specify posting interval (minutes between slots, 0 = queue/blank time)
    python upload_clips_to_b2.py --interval 60

Requirements
------------
    pip install boto3 openpyxl python-dotenv

Credentials are read from the project .env file (or environment variables):
    B2_KEY_ID, B2_APPLICATION_KEY, B2_BUCKET_NAME, B2_ENDPOINT_URL
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Bootstrap: ensure the factory root is on sys.path so imports resolve
# ---------------------------------------------------------------------------
_FACTORY_ROOT = Path(__file__).parent.resolve()
if str(_FACTORY_ROOT) not in sys.path:
    sys.path.insert(0, str(_FACTORY_ROOT))

# Load .env from the factory root
try:
    from dotenv import load_dotenv
    load_dotenv(_FACTORY_ROOT / ".env", override=False)
except ImportError:
    pass  # dotenv optional — rely on shell environment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("upload_clips_to_b2")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
_DEFAULT_PAGE    = "wonder_feed"
_DEFAULT_CLIPS   = "outputs/{page}/clips"
_VIDEO_EXTS      = {".mp4", ".mov", ".mkv", ".webm"}
_DEFAULT_CAPTION = (
    "A profound exploration of emotional intelligence and relationship dynamics. "
    "Follow for more psychology-driven insights. 💙 #Relationships #EmotionalIntelligence"
)


# ---------------------------------------------------------------------------
# B2 upload helpers (inline — works without avatar_engine on PATH)
# ---------------------------------------------------------------------------

def _build_b2_resource():
    try:
        import boto3
        from botocore.config import Config
    except ImportError:
        raise RuntimeError("boto3 not installed. Run: pip install boto3")

    key_id  = (os.getenv("B2_KEY_ID") or "").strip()
    app_key = (os.getenv("B2_APPLICATION_KEY") or "").strip()
    if not key_id or not app_key:
        raise RuntimeError(
            "B2_KEY_ID and B2_APPLICATION_KEY are not set.\n"
            "Add them to your .env file or export them in the shell."
        )
    endpoint = (
        os.getenv("B2_ENDPOINT_URL") or "https://s3.us-east-005.backblazeb2.com"
    ).strip()

    return boto3.resource(
        service_name="s3",
        endpoint_url=endpoint,
        aws_access_key_id=key_id,
        aws_secret_access_key=app_key,
        region_name="us-east-005",
        config=Config(
            signature_version="s3v4",
            s3={"addressing_style": "path"},
        ),
    )


def _object_exists(b2, bucket: str, key: str) -> bool:
    try:
        b2.Object(bucket, key).load()
        return True
    except Exception as exc:
        try:
            code = exc.response["Error"]["Code"]
            if code in ("404", "NoSuchKey", "403"):
                return False
        except (AttributeError, KeyError):
            pass
        return False


def _public_url(bucket: str, filename: str) -> str:
    return f"https://{bucket}.s3.us-east-005.backblazeb2.com/{Path(filename).name}"


def upload_file_to_b2(
    b2,
    bucket: str,
    local_path: Path,
    *,
    dry_run: bool = False,
) -> str:
    """Upload one file; return its public URL."""
    key = local_path.name
    url = _public_url(bucket, key)

    if dry_run:
        print(f"  [DRY-RUN] Would upload: {key} → {url}")
        return url

    if _object_exists(b2, bucket, key):
        print(f"  [SKIP] Already in B2: {key}")
        logger.info("Already uploaded: %s", key)
        return url

    size_mb = local_path.stat().st_size / (1024 * 1024)
    print(f"  [UPLOAD] {key} ({size_mb:.1f} MB) …")
    b2.Object(bucket, key).upload_file(
        str(local_path),
        ExtraArgs={"ContentType": "video/mp4"},
    )
    print(f"  [OK] {url}")
    logger.info("Uploaded: %s → %s", key, url)
    return url


# ---------------------------------------------------------------------------
# PostPlanner XLSX writer
# ---------------------------------------------------------------------------

def _posting_slots(
    count: int,
    interval_minutes: int,
    start: datetime | None = None,
) -> list[str]:
    """Generate posting time strings (MM/DD/YYYY HH:MM) or blanks."""
    if interval_minutes <= 0:
        return [""] * count
    base = start or datetime.now(timezone.utc).replace(
        hour=9, minute=0, second=0, microsecond=0
    )
    return [
        (base + timedelta(minutes=interval_minutes * i)).strftime("%m/%d/%Y %H:%M")
        for i in range(count)
    ]


def _last_data_row(ws) -> int:
    """
    Return the index of the last row that contains at least one non-empty cell.

    openpyxl's ``ws.max_row`` counts every row the XML parser has seen,
    including empty placeholder rows that templates bake in as spacers.
    Walking backward from ``max_row`` until we hit a non-blank cell gives
    us the true last occupied row so our data starts immediately after it.
    """
    for row_idx in range(ws.max_row, 0, -1):
        if any(
            cell.value not in (None, "")
            for cell in ws[row_idx]
        ):
            return row_idx
    return 0  # completely empty sheet


def write_postplanner_xlsx(
    records: list[dict],
    output_path: Path,
    *,
    template_path: Path | None = None,
) -> Path:
    """
    Write a PostPlanner XLSX matching the 3-column schema:
        A: DATE / TIME  |  B: CAPTION  |  C: MEDIA URL

    When a template is supplied the workbook is cloned and data is written
    immediately below the last non-empty row (header block / instructions).
    Empty placeholder rows that templates bake in are skipped so there is
    no blank gap between the header and the first data row.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["DATE / TIME", "CAPTION", "MEDIA URL"]

    if template_path and template_path.is_file():
        wb = load_workbook(str(template_path))
        ws = wb.active
        # Find the real last occupied row — not the XML-inflated max_row.
        last_real = _last_data_row(ws)
        start_row = last_real + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PostPlanner"
        ws.append(headers)
        start_row = 2

    for i, rec in enumerate(records):
        row = start_row + i
        ws.cell(row=row, column=1, value=rec.get("posting_time", ""))
        ws.cell(row=row, column=2, value=rec.get("caption", ""))
        ws.cell(row=row, column=3, value=rec.get("media_url", ""))

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Upload existing video clips to Backblaze B2 and generate a PostPlanner XLSX.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--clips-dir",
        help="Directory containing .mp4 clips to upload. "
             "Defaults to outputs/<page>/clips/",
    )
    parser.add_argument(
        "--page",
        default=_DEFAULT_PAGE,
        help=f"Page ID (default: {_DEFAULT_PAGE})",
    )
    parser.add_argument(
        "--output",
        help="Path for the generated postplanner XLSX. "
             "Defaults to outputs/<page>/postplanner/PostPlanner_B2_<timestamp>.xlsx",
    )
    parser.add_argument(
        "--caption",
        default="",
        help="Default caption applied to every row. "
             "If not provided, a fallback caption is used.",
    )
    parser.add_argument(
        "--captions-json",
        help="Path to a JSON file mapping filename → caption string. "
             "Overrides --caption on a per-file basis.",
    )
    parser.add_argument(
        "--interval",
        type=int,
        default=60,
        metavar="MINUTES",
        help="Minutes between posting slots (0 = blank/queue mode). Default: 60",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be uploaded without touching B2 or writing files.",
    )
    parser.add_argument(
        "--bucket",
        help="Override B2_BUCKET_NAME env var.",
    )
    args = parser.parse_args(argv)

    # ── Resolve paths ─────────────────────────────────────────────────────────
    clips_dir = Path(
        args.clips_dir
        or _FACTORY_ROOT / _DEFAULT_CLIPS.format(page=args.page)
    )
    if not clips_dir.is_dir():
        print(f"[ERROR] Clips directory not found: {clips_dir}")
        sys.exit(1)

    clips = sorted(
        p for p in clips_dir.iterdir()
        if p.is_file() and p.suffix.lower() in _VIDEO_EXTS
    )
    if not clips:
        print(f"[ERROR] No video files found in: {clips_dir}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  Clips directory : {clips_dir}")
    print(f"  Video files     : {len(clips)}")
    print(f"  Dry-run         : {args.dry_run}")
    print(f"{'='*60}\n")

    # ── Per-file captions ─────────────────────────────────────────────────────
    caption_map: dict[str, str] = {}
    if args.captions_json:
        try:
            caption_map = json.loads(Path(args.captions_json).read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"[WARN] Could not load captions JSON ({exc}); using default caption.")

    default_caption = args.caption or _DEFAULT_CAPTION

    # ── B2 connection ─────────────────────────────────────────────────────────
    bucket = (args.bucket or os.getenv("B2_BUCKET_NAME") or "MediaupscaleStorage").strip()
    b2 = None
    if not args.dry_run:
        try:
            b2 = _build_b2_resource()
        except RuntimeError as exc:
            print(f"[ERROR] {exc}")
            sys.exit(1)

    # ── Upload loop ───────────────────────────────────────────────────────────
    slots = _posting_slots(len(clips), args.interval)
    records: list[dict] = []

    for clip, slot in zip(clips, slots):
        caption = caption_map.get(clip.name) or caption_map.get(clip.stem) or default_caption
        try:
            url = upload_file_to_b2(b2, bucket, clip, dry_run=args.dry_run)
        except Exception as exc:
            print(f"  [FAIL] {clip.name}: {exc}")
            logger.error("Upload failed for %s: %s", clip.name, exc)
            url = ""

        records.append({
            "posting_time": slot,
            "caption": caption,
            "media_url": url,
        })

    # ── Write postplanner XLSX ────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(
        args.output
        or _FACTORY_ROOT
        / f"outputs/{args.page}/postplanner/PostPlanner_B2_{timestamp}.xlsx"
    )

    template_candidates = [
        _FACTORY_ROOT / "sample_bulk_posts_import_3.xlsx",
        _FACTORY_ROOT / "sample_bulk_posts_import.xlsx",
    ]
    template = next((t for t in template_candidates if t.is_file()), None)

    if args.dry_run:
        print(f"\n[DRY-RUN] Would write PostPlanner XLSX → {output_path}")
        print(f"[DRY-RUN] {len(records)} rows | template: {template or 'none'}\n")
        for r in records:
            print(f"  {r['posting_time'] or '(queue)'}  |  {r['media_url']}")
    else:
        out = write_postplanner_xlsx(records, output_path, template_path=template)
        print(f"\n[DONE] PostPlanner written → {out}")
        print(f"       {len(records)} rows | {sum(1 for r in records if r['media_url'])} with B2 URLs\n")


if __name__ == "__main__":
    main()
