"""fix_uploaded_titles.py — Retroactive B2 upload + Excel MEDIA URL patcher.

Scans every reel MP4 under ``outputs/ancient_knowledge/``, uploads it to
Backblaze B2 (idempotent — files already there are skipped), then patches
the MEDIA URL column in every postplan_*.xlsx and automated_bulk_posts_import.xlsx
to contain the live .mp4 URL instead of an ImgBB image link.

Usage
-----
    python fix_uploaded_titles.py                    # ancient_knowledge (default)
    python fix_uploaded_titles.py --page wonder_feed
    python fix_uploaded_titles.py --dry-run          # preview only, no writes

Matching strategy
-----------------
1. Load ``content_library.json`` to get exact (caption → video_path) pairs.
2. For each xlsx row, compute the longest word-overlap between the row caption
   and each library entry's caption.  Best match above a threshold wins.
3. If the matched entry has a video_path that exists on disk, upload to B2
   and write the resulting URL into the MEDIA URL column.
4. Rows that already contain a B2-style URL are skipped.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

try:
    from dotenv import load_dotenv
    load_dotenv(_SCRIPT_DIR / ".env", override=False)
except ImportError:
    pass

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
_LOG = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
_B2_URL_PREFIX = "https://"
_IMGBB_PREFIX  = "https://i.ibb.co"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _slug_words(text: str) -> set[str]:
    """Return lower-case word tokens from *text* (≥ 3 chars)."""
    return {w.lower() for w in re.findall(r"[a-zA-Z]{3,}", text)}


def _is_b2_url(url: str) -> bool:
    """Return True when *url* looks like a Backblaze B2 public URL."""
    return bool(url) and "backblazeb2.com" in url


def _load_library(page: str) -> list[dict[str, Any]]:
    lib_path = _SCRIPT_DIR / "outputs" / page / "content_library.json"
    if not lib_path.is_file():
        _LOG.warning("content_library.json not found at %s", lib_path)
        return []
    try:
        data = json.loads(lib_path.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception as exc:
        _LOG.warning("Failed to load content_library.json: %s", exc)
        return []


def _find_xlsx_files(page: str) -> list[Path]:
    """Return all postplan_*.xlsx and automated_bulk_posts_import.xlsx for *page*."""
    out_dir = _SCRIPT_DIR / "outputs" / page
    if not out_dir.is_dir():
        return []
    xlsx_files: list[Path] = []
    # automated_bulk_posts_import.xlsx
    bulk = out_dir / "automated_bulk_posts_import.xlsx"
    if bulk.is_file():
        xlsx_files.append(bulk)
    # postplanner sub-directory
    pp_dir = out_dir / "postplanner"
    if pp_dir.is_dir():
        xlsx_files.extend(sorted(pp_dir.glob("postplan_*.xlsx")))
    # also look directly in out_dir for any stray postplan files
    xlsx_files.extend(sorted(out_dir.glob("postplan_*.xlsx")))
    # deduplicate while preserving order
    seen: set[Path] = set()
    unique: list[Path] = []
    for p in xlsx_files:
        if p not in seen:
            seen.add(p)
            unique.append(p)
    return unique


def _best_library_match(
    row_caption: str,
    library: list[dict[str, Any]],
    min_overlap: int = 4,
) -> dict[str, Any] | None:
    """Return the library entry whose caption has the best word-overlap with *row_caption*."""
    row_words = _slug_words(row_caption)
    if not row_words:
        return None

    best_entry: dict | None = None
    best_score = 0

    for entry in library:
        lib_cap  = entry.get("final_caption") or entry.get("caption") or ""
        lib_words = _slug_words(lib_cap)
        overlap   = len(row_words & lib_words)
        if overlap > best_score:
            best_score  = overlap
            best_entry  = entry

    if best_score >= min_overlap:
        return best_entry
    return None


# ---------------------------------------------------------------------------
# B2 upload
# ---------------------------------------------------------------------------

def _upload_to_b2(video_path: Path, dry_run: bool) -> str:
    """Upload *video_path* to B2 and return its public URL.

    Uses ``B2VideoUploader`` which skips re-upload if the object already exists.
    Returns an empty string if the upload fails or dry_run is True.
    """
    if dry_run:
        from avatar_engine.b2_client import B2VideoUploader
        return B2VideoUploader.public_url(video_path.name)

    try:
        from avatar_engine.b2_client import B2VideoUploader
        uploader = B2VideoUploader()
        url = uploader.upload(video_path)
        _LOG.info("B2 OK  %s -> %s", video_path.name, url)
        return url
    except Exception as exc:
        _LOG.error("B2 upload failed for %s: %s", video_path.name, exc)
        return ""


# ---------------------------------------------------------------------------
# Excel patcher
# ---------------------------------------------------------------------------

def _patch_xlsx(
    xlsx_path: Path,
    caption_to_b2: dict[str, str],   # caption_snippet → B2 URL
    dry_run: bool,
) -> int:
    """Patch MEDIA URL column in *xlsx_path* using *caption_to_b2*.

    Returns number of rows updated.
    """
    from openpyxl import load_workbook  # type: ignore[import]

    try:
        wb = load_workbook(xlsx_path)
    except Exception as exc:
        _LOG.warning("Cannot open %s: %s - skipping.", xlsx_path.name, exc)
        return 0

    updated = 0
    try:
        for ws in wb.worksheets:
            # Detect column indices from header row
            col_caption = col_media = None
            header_row = ws[1]
            for cell in header_row:
                v = str(cell.value or "").strip().upper()
                if "CAPTION" in v:
                    col_caption = cell.column
                elif "MEDIA" in v or "URL" in v:
                    col_media = cell.column

            # Fall back to column 2=caption, 3=media (standard layout)
            if col_caption is None:
                col_caption = 2
            if col_media is None:
                col_media = 3

            for row in ws.iter_rows(min_row=2):
                if len(row) < max(col_caption, col_media):
                    continue
                caption_cell = row[col_caption - 1]
                media_cell   = row[col_media   - 1]

                caption_text = str(caption_cell.value or "")
                media_val    = str(media_cell.value   or "")

                # Skip rows that already have a B2 URL
                if _is_b2_url(media_val):
                    continue

                # Find the best matching B2 URL by caption word-overlap
                cap_words = _slug_words(caption_text)
                best_url  = ""
                best_score = 0

                for snippet, b2_url in caption_to_b2.items():
                    overlap = len(cap_words & _slug_words(snippet))
                    if overlap > best_score:
                        best_score = overlap
                        best_url   = b2_url

                if best_url and best_score >= 4:
                    _LOG.info(
                        "  [%s] row %s: MEDIA URL -> %s  (overlap=%d)",
                        xlsx_path.name, caption_cell.row, best_url[:70], best_score,
                    )
                    if not dry_run:
                        media_cell.value = best_url
                    updated += 1

        if updated and not dry_run:
            wb.save(xlsx_path)
            _LOG.info("Saved %s (%d rows updated)", xlsx_path.name, updated)
    finally:
        if callable(getattr(wb, "close", None)):
            wb.close()

    return updated


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def run(page: str, dry_run: bool) -> None:
    marker = "[DRY RUN] " if dry_run else ""
    print(f"\n{'='*70}")
    print(f"  {marker}B2 Upload + Excel MEDIA URL Patcher")
    print(f"  Page: {page}")
    print(f"{'='*70}")

    # ── Step 1: collect all reel MP4s ─────────────────────────────────────
    out_dir = _SCRIPT_DIR / "outputs" / page
    if not out_dir.is_dir():
        _LOG.error("Output directory not found: %s", out_dir)
        sys.exit(1)

    mp4_files = sorted(out_dir.rglob("reel_*.mp4"))
    if not mp4_files:
        _LOG.warning("No reel_*.mp4 files found under %s", out_dir)
        return

    print(f"\nFound {len(mp4_files)} reel MP4 file(s):")
    for p in mp4_files:
        mb = p.stat().st_size / 1_048_576
        print(f"  * {p.name} ({mb:.1f} MB)")

    # ── Step 2: upload each MP4 to B2 ────────────────────────────────────
    print(f"\n{marker}Uploading to Backblaze B2 ...")
    mp4_to_b2: dict[str, str] = {}   # filename → B2 URL
    for mp4 in mp4_files:
        url = _upload_to_b2(mp4, dry_run)
        if url:
            mp4_to_b2[mp4.name] = url
        else:
            _LOG.warning("No B2 URL for %s - it will be skipped in Excel.", mp4.name)

    if not mp4_to_b2:
        _LOG.error("No B2 URLs generated - check B2 credentials and bucket config.")
        sys.exit(1)

    # ── Step 3: build caption → B2 URL lookup via content_library.json ───
    library = _load_library(page)
    caption_to_b2: dict[str, str] = {}

    for entry in library:
        vp_str = entry.get("video_path") or ""
        if not vp_str:
            continue
        vp_name = Path(vp_str).name
        b2_url  = mp4_to_b2.get(vp_name)
        if b2_url:
            cap = entry.get("final_caption") or entry.get("caption") or ""
            if cap:
                caption_to_b2[cap] = b2_url

    # Fallback: for any MP4 without a library entry, add by filename-derived slug
    for fname, b2_url in mp4_to_b2.items():
        if not any(Path(k).name == fname for k in caption_to_b2 if k.startswith("reel_")):
            slug = re.sub(r"_v\d+\.mp4$", "", fname).replace("reel_", "").replace("_", " ")
            caption_to_b2[slug] = b2_url

    print(f"\nCaption->B2 mapping built for {len(caption_to_b2)} video(s).")

    # ── Step 4: patch all xlsx files ──────────────────────────────────────
    xlsx_files = _find_xlsx_files(page)
    if not xlsx_files:
        _LOG.warning("No xlsx files found for page '%s'.", page)
        return

    print(f"\n{marker}Patching {len(xlsx_files)} xlsx file(s) ...")
    total_updated = 0
    for xf in xlsx_files:
        n = _patch_xlsx(xf, caption_to_b2, dry_run)
        total_updated += n
        if n == 0:
            print(f"  (no changes) {xf.name}")
        else:
            status = "would update" if dry_run else "updated"
            print(f"  OK {status} {n} row(s) -> {xf.name}")

    print(f"\n{'='*70}")
    verb = "Would have updated" if dry_run else "Updated"
    print(f"  {verb} {total_updated} cell(s) across {len(xlsx_files)} file(s).")
    if dry_run:
        print("  Run without --dry-run to apply changes.")
    print(f"{'='*70}\n")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Upload reel MP4s to B2 and patch MEDIA URL in all xlsx files."
    )
    ap.add_argument("--page", default="ancient_knowledge",
                    help="Page slug (default: ancient_knowledge).")
    ap.add_argument("--dry-run", "-n", action="store_true",
                    help="Preview changes without writing anything.")
    args = ap.parse_args()
    run(args.page, args.dry_run)


if __name__ == "__main__":
    main()
